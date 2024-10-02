import re
import os
import sys
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers



def validar_data(startDate, endDate):
    try:
        # Converte as strings de data em objetos datetime
        start_date_obj = datetime.strptime(startDate, "%Y-%m-%d")
        end_date_obj = datetime.strptime(endDate, "%Y-%m-%d")

        # Verifica se startDate é anterior a endDate
        if start_date_obj <= end_date_obj:
            return True
        else:
            print("A data de início é posterior à data de término.")
            return False
    except ValueError as e:
        print(f"Erro na validação das datas: {e}")
        return False


def normalizar_texto(texto):
    """Remove caracteres não alfanuméricos e converte o texto para minúsculas."""
    return re.sub(r'\W+', '', texto).lower()


def clear_screen():
    """ Limpa o terminal de acordo com o sistema operacional. """
    if sys.platform == 'win32':
        os.system('cls')  # Para Windows
    else:
        os.system('clear')  # Para Unix/Linux


def formata_cpj(value):
    if pd.notna(value):
        # Tentar converter para inteiro e depois para string
        try:
            # Remove espaços e converte para um número inteiro antes de converter para string
            return str(int(float(value))).zfill(14)
        except ValueError:
            return value  # Retorna o valor original se a conversão falhar
    return value



def formata_para_real(excel_filename, columns_to_format):
    # Carregar o arquivo Excel para aplicar a formatação
    wb = load_workbook(excel_filename)
    ws = wb.active

    # Formato de moeda brasileiro
    currency_format = 'R$ #,##0.00'

    # Aplica a formatação de moeda em cada célula das colunas especificadas
    for col in columns_to_format:
        for cell in ws[col][1:]:  # Pule o cabeçalho
            cell.number_format = currency_format

    # Salva o arquivo com a formatação aplicada
    wb.save(excel_filename)


