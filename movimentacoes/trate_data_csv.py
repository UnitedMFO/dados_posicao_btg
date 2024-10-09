import pandas as pd
import os
from openpyxl import load_workbook
from utils import formata_cpj, formata_para_real
from openpyxl.styles import PatternFill



def recebe_e_cria_movimentacao(csv_filename, base_nome_arquivo):
    # Define o diretório onde o arquivo Excel será salvo
    diretorio_destino = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'relatorios')
    os.makedirs(diretorio_destino, exist_ok=True)  # Cria o diretório "relatorios" se ele não existir

    # Lê o arquivo CSV
    df = pd.read_csv(csv_filename, delimiter=',', encoding='utf-8', header=0)

    # Define as colunas esperadas
    expected_columns = ['marketName', 'subMarketName', 'asset', 'fundCnpj', 'movementDate',
                        'movementHistory', 'launchType', 'grossValue', 'irValue', 'iofValue',
                        'dueDate', 'index', 'fee', 'issuer', 'accountingGroupCode']

    # Filtra o DataFrame para manter apenas as colunas esperadas
    df = df[expected_columns]

    # Cria o nome do arquivo Excel e evita sobrescrever arquivos existentes
    contador = 1
    nome_arquivo = base_nome_arquivo
    caminho_arquivo = os.path.join(diretorio_destino, f"{nome_arquivo}.xlsx")

    while os.path.exists(caminho_arquivo):
        nome_arquivo = f"{base_nome_arquivo} ({contador})"
        caminho_arquivo = os.path.join(diretorio_destino, f"{nome_arquivo}.xlsx")
        contador += 1

    # Salva o DataFrame como Excel no diretório correto
    df.to_excel(caminho_arquivo, index=False)

    # Remove o arquivo CSV original
    os.remove(csv_filename)
    print(f"Arquivo Excel criado com sucesso")

    return caminho_arquivo


def formata_movimentacao(excel_filename):
    list_negativos = ['CRÉDITO', 'JUROS', 'JUROS S/ CAPITAL', 'RECEBIMENTO DIVIDENDOS', 'RI', 'RS', 'VENCIMENTO DE TÍTULO', 'VENDA']
    excel_df = pd.read_excel(excel_filename)

    # Formata o CNPJ na coluna 'fundCnpj'
    excel_df['fundCnpj'] = excel_df['fundCnpj'].apply(formata_cpj)

    # Percorre cada linha do DataFrame
    for index, row in excel_df.iterrows():
        # Verifica se 'fundCnpj' está vazio e aplica as lógicas para CC, RF, ou ACOES
        if pd.isna(row['fundCnpj']):
            if row['subMarketName'] == 'CC':
                excel_df.at[index, 'fundCnpj'] = 'cash'
            elif row['subMarketName'] in ['RF', 'ACOES']:
                excel_df.at[index, 'fundCnpj'] = row['asset']

        # Verifica se o valor de 'launchType' contém qualquer item da lista_negativos
        # Se corresponder, torna o valor de 'grossValue' negativo
        if any(neg in str(row['launchType']).upper() for neg in list_negativos):
            excel_df.at[index, 'grossValue'] = -abs(row['grossValue'])

    # Salvando as mudanças de volta no arquivo Excel
    excel_df.to_excel(excel_filename, index=False)
    print(f"Formatação concluída com sucesso")



def calcula_valor_liquido(excel_filename):
    # Lê o arquivo Excel
    excel_df = pd.read_excel(excel_filename)

    # Calcula o valor líquido, considerando casos onde 'irValue' ou 'iofValue' podem estar vazios
    excel_df['netValue'] = excel_df.apply(
        lambda row: row['grossValue'] if pd.isna(row['irValue']) or pd.isna(row['iofValue'])
        else row['grossValue'] - row['irValue'] - row['iofValue'],
        axis=1
    )

    # Inserir a coluna 'netValue' logo após 'iofValue'
    cols = excel_df.columns.tolist()  # Lista atual de colunas
    iof_index = cols.index('iofValue')  # Encontrar o índice de 'iofValue'
    # Inserir 'netValue' logo após 'iofValue'
    cols.insert(iof_index + 1, cols.pop(cols.index('netValue')))
    excel_df = excel_df[cols]  # Reordenar o DataFrame com a nova ordem de colunas

    # Salvando as mudanças de volta no arquivo Excel
    excel_df.to_excel(excel_filename, index=False)
    formata_para_real(excel_filename, ['H', 'I', 'J', 'K'])
    print(f"Valor líquido calculado com sucesso")





def destaca_negativos(excel_filename):
    # Carrega o arquivo Excel para aplicar a formatação com openpyxl
    wb = load_workbook(excel_filename)
    ws = wb.active

    # Define um preenchimento amarelo para células negativas
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Aplica o preenchimento amarelo nas células da coluna 'grossValue' com valores negativos
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=8, max_col=8):  # Coluna H é a 8ª
        for cell in row:
            if cell.value is not None and cell.value < 0:
                cell.fill = yellow_fill

    # Salva o arquivo Excel com as formatações aplicadas
    wb.save(excel_filename)






