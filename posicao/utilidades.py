import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
from datetime import datetime, date, timedelta


def obter_data_mes_anterior(data_atual):
    """Retorna a data do mês anterior ou permite ao usuário inserir uma data manualmente."""
    while True:
        opcao = input(f"Rentabilidade do fundo na data anterior a ({data_atual}) ou inserir manualmente (1 - manual / 0 - mês anterior)? ")

        if opcao == '1':
            while True:
                data_manual = input("Digite a data desejada (AAAA-MM-DD): ")
                if validar_formato_data(data_manual):
                    return data_manual
                else:
                    print("Formato de data inválido, tente novamente.")

        elif opcao == '0':
            ano, mes, dia = map(int, data_atual.split('-'))

            if mes == 1:
                mes = 12
                ano -= 1
            else:
                mes -= 1

            return f"{ano}-{mes:02d}-{dia}"

        else:
            print("Opção inválida. Escolha 0 para mês anterior ou 1 para inserir manualmente.")


def obter_meses_atual_e_anterior():
    """Retorna o mês atual e o anterior no formato MM/AAAA."""
    data_atual = date.today()
    ano_atual = data_atual.year
    mes_atual = data_atual.month

    if mes_atual == 1:
        mes_anterior = 12
        ano_anterior = ano_atual - 1
    else:
        mes_anterior = mes_atual - 1
        ano_anterior = ano_atual

    mes_atual_formatado = f'{mes_atual:02d}/{ano_atual}'
    mes_anterior_formatado = f'{mes_anterior:02d}/{ano_anterior}'

    return mes_anterior_formatado, mes_atual_formatado


def obter_dois_meses_anteriores():
    """Retorna os dois meses anteriores ao atual no formato MM/AAAA."""
    data_atual = date.today()
    ano_atual = data_atual.year
    mes_atual = data_atual.month

    if mes_atual == 1:
        mes_anterior = 12
        mes_anterior_anterior = 11
        ano_anterior = ano_atual - 1
    elif mes_atual == 2:
        mes_anterior = 1
        mes_anterior_anterior = 12
        ano_anterior = ano_atual
        ano_anterior_anterior = ano_atual - 1
    else:
        mes_anterior = mes_atual - 1
        mes_anterior_anterior = mes_atual - 2
        ano_anterior = ano_atual
        ano_anterior_anterior = ano_atual

    mes_anterior_formatado = f'{mes_anterior:02d}/{ano_anterior}'
    mes_anterior_anterior_formatado = f'{mes_anterior_anterior:02d}/{ano_anterior_anterior}'

    return mes_anterior_anterior_formatado, mes_anterior_formatado


def validar_formato_data(data_entrada):
    """Valida o formato da data fornecida no formato AAAA-MM-DD."""
    try:
        datetime.strptime(data_entrada, "%Y-%m-%d")
        return True
    except ValueError:
        return False


def limpar_tela():
    """Limpa o terminal de acordo com o sistema operacional."""
    if sys.platform == 'win32':
        os.system('cls')
    else:
        os.system('clear')


def aplicar_formatacao_excel(caminho_arquivo_excel):
    """Aplica formatação monetária e ajusta larguras de colunas em um arquivo Excel."""
    try:
        workbook = load_workbook(caminho_arquivo_excel)

        for nome_planilha in workbook.sheetnames:
            planilha = workbook[nome_planilha]

            # Formatar cabeçalhos: negrito, fundo amarelo e alinhamento central
            for celula in planilha[1]:  # Primeira linha (cabeçalhos)
                celula.font = Font(bold=True)
                celula.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                celula.alignment = Alignment(horizontal='center', vertical='center')

            # Ajustar largura das colunas
            for coluna in planilha.columns:
                largura_maxima = 0
                letra_coluna = coluna[0].column_letter
                for celula in coluna:
                    if celula.value:
                        largura_maxima = max(largura_maxima, len(str(celula.value)) + 2)
                planilha.column_dimensions[letra_coluna].width = largura_maxima * 1.2

            # Aplicar formatação monetária nas células numéricas
            for linha in planilha.iter_rows(min_row=2, max_row=planilha.max_row, min_col=1, max_col=planilha.max_column):
                for celula in linha:
                    if isinstance(celula.value, (float, int)):
                        celula.number_format = 'R$ #,##0.00'

        workbook.save(caminho_arquivo_excel)

    except Exception as e:
        print(f"Erro ao aplicar formatação: {str(e)}")
