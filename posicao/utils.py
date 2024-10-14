import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
from datetime import datetime,date,timedelta


def mes_anterior_fundo(data):
    while True:
        # Solicitar ao usuário se deseja a data de um mês anterior ou inserir manualmente
        opcao = input(f"Rentabilidade do fundo a data anterior a ({data}) ou inserir uma data manualmente (1 - manual / 0 - mes anterior) ? ")

        if opcao == '1':
            # Pergunta para o usuário inserir a data manualmente
            while True:
                data_manual = input("Digite a data desejada (AAAA-MM-DD): ")
                if validar_data(data_manual):
                    return data_manual
                else:
                    print("Formato de data inválido, tente novamente.")

        elif opcao == '0':
            # Se a opção for automática, calcular o mês anterior
            data_atual = data.split('-')
            ano = int(data_atual[0])
            mes = int(data_atual[1])
            dia = data_atual[2]

            # Se o mês for janeiro, volta para dezembro do ano anterior
            if mes == 1:
                mes = 12
                ano -= 1
            else:
                mes -= 1

            # Formatar mês com dois dígitos
            mes_formatado = f"{mes:02d}"

            # Retornar a data no formato yyyy-mm-dd
            date_req = f"{ano}-{mes_formatado}-{dia}"
            return date_req

        else:
            print("Opção inválida. Escolha 0 para pegar o mês anterior ou 1 para inserir a data manualmente.")


def data_atual():
    # Obtém a data atual e separa em ['YYYY', 'MM', 'DD']
    data_atual = str(date.today()).split('-')

    # Extrai o mês e o ano atuais
    ano_atual = int(data_atual[0])
    mes_atual = int(data_atual[1])

    # Calcula o mês anterior e ajusta o ano se necessário
    if mes_atual == 1:
        # Se for janeiro, o mês anterior é dezembro do ano anterior
        mes_anterior = 12
        ano_anterior = ano_atual - 1
    else:
        mes_anterior = mes_atual - 1
        ano_anterior = ano_atual

    # Formata os meses com dois dígitos
    mes_atual_formatado = f'{mes_atual:02d}/{ano_atual}'
    mes_anterior_formatado = f'{mes_anterior:02d}/{ano_anterior}'

    # Retorna os dois meses formatados
    return mes_anterior_formatado, mes_atual_formatado



def data_mes_anterior():
    # Data de teste (você pode substituir pelo uso de `date.today()` se necessário)
    data_atual = str(date.today()).split('-')

    # Extrai o mês e ano atuais
    ano_atual = int(data_atual[0])
    mes_atual = int(data_atual[1])

    # Calcula o mês anterior e o mês anterior ao anterior
    if mes_atual == 1:
        # Caso seja janeiro, o mês anterior é dezembro do ano anterior, e o anterior a esse é novembro do mesmo ano anterior
        mes_anterior = 12
        mes_anterior_anterior = 11
        ano_anterior = ano_atual - 1
        ano_anterior_anterior = ano_atual - 1
    elif mes_atual == 2:
        # Caso seja fevereiro, o mês anterior é janeiro, e o anterior a esse é dezembro do ano anterior
        mes_anterior = 1
        mes_anterior_anterior = 12
        ano_anterior = ano_atual
        ano_anterior_anterior = ano_atual - 1
    else:
        # Para todos os outros meses
        mes_anterior = mes_atual - 1
        mes_anterior_anterior = mes_atual - 2
        ano_anterior = ano_atual
        ano_anterior_anterior = ano_atual

    # Formata os meses com dois dígitos
    mes_anterior_formatado = f'{mes_anterior:02d}/{ano_anterior}'
    mes_anterior_anterior_formatado = f'{mes_anterior_anterior:02d}/{ano_anterior_anterior}'

    # Retorna os dois meses formatados
    return mes_anterior_anterior_formatado, mes_anterior_formatado


def validar_data(data_entrada):
    try:
        datetime.strptime(data_entrada, "%Y-%m-%d")
        return True
    except ValueError:
        return False



def clear_screen():
    """ Limpa o terminal de acordo com o sistema operacional. """
    if sys.platform == 'win32':
        os.system('cls')  # Para Windows
    else:
        os.system('clear')  # Para Unix/Linux


def aplicar_formatacao_monetaria(workbook_path):
    try:
        wb = load_workbook(workbook_path)

        for sheetname in wb.sheetnames:
            ws = wb[sheetname]

            # Formatar cabeçalhos: negrito, fundo amarelo e alinhamento central
            for cell in ws[1]:  # ws[1] refere-se à primeira linha, que geralmente é a linha dos cabeçalhos
                cell.font = Font(bold=True)  # Negrito
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Fundo amarelo
                cell.alignment = Alignment(horizontal='center', vertical='center')  # Alinhamento centralizado

            # Ajusta a largura das colunas para se adequar ao conteúdo
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter  # Obtém a letra da coluna (ex: 'A', 'B', etc.)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)) + 2)  # Adiciona espaço extra
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2  # Ajusta a largura com um pequeno incremento
                ws.column_dimensions[column_letter].width = adjusted_width

            # Percorre todas as células da planilha
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    # Aplica formatação de moeda às células com valores numéricos
                    if isinstance(cell.value, (float, int)):
                        cell.number_format = 'R$ #,##0.00'

        wb.save(workbook_path)

    except Exception as e:
        print(f"Erro ao aplicar formatação: {str(e)}")


